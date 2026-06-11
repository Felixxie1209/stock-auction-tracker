import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from datetime import datetime

# Data
auction_data = [
    {
        "id": 1, "name": "思创智联", "code": "300078", "market": "SZ",
        "shares": "思创医惠集团持有302.89万股（分2笔）",
        "platform": "阿里拍卖", "court": "浙江省杭州市中级人民法院",
        "date": "2026-06-08", "date_end": "2026-06-09",
        "date_display": "06-08 10:00", "date_end_display": "06-09 10:00",
        "price": "190.92万 + 682.63万", "price_note": "企业破产清算",
        "summary": "破产管理人分2笔处置思创医惠集团持有的上市公司股票：66.2万股(起拍190.92万)和236.69万股(起拍682.63万)，均为无限售流通股。",
        "background": "思创智联(原思创医惠)主营智慧医疗和物联网，集团公司进入破产清算程序，其持有的上市公司股份被分拆拍卖处置。两笔标的同日开拍。",
        "status": "进行中"
    },
    {
        "id": 2, "name": "群兴玩具", "code": "002575", "market": "SZ",
        "shares": "深圳星河持有140万股（第二次拍卖）",
        "platform": "阿里拍卖", "court": "广东省深圳市龙岗区人民法院",
        "date": "2026-06-08", "date_end": "2026-06-09",
        "date_display": "06-08 10:00", "date_end_display": "06-09 10:00",
        "price": "762.72万元", "price_note": "评估价1153.6万×70%",
        "summary": "二次拍卖群兴玩具140万股无限售流通股，评估价1153.6万元，起拍价762.72万元(评估价70%)，保证金76.27万元。",
        "background": "群兴玩具第一大股东深圳星河因债务纠纷被强制执行。一拍流拍后进入二拍程序，起拍价按评估价70%确定。累计拟拍卖约970万股，控制权面临不确定性。",
        "status": "进行中"
    },
    {
        "id": 3, "name": "国新证券", "code": "非上市", "market": "",
        "shares": "科铭实业持有4000万股（原华融证券）",
        "platform": "阿里拍卖", "court": "广东省深圳市福田区人民法院",
        "date": "2026-06-11", "date_end": "2026-06-12",
        "date_display": "06-11 10:00", "date_end_display": "06-12 10:00",
        "price": "8652万元", "price_note": "评估价1.236亿×70%",
        "summary": "拍卖科铭实业持有的国新证券(原华融证券)4000万股，评估价1.236亿元，起拍价8652万元，保证金1290万元。",
        "background": "国新证券(原华融证券)为非上市全国性券商。科铭实业因涉工行债务纠纷被深圳福田区法院强制执行，其持有的券商股份被裁定拍卖。",
        "status": "2天后"
    },
    {
        "id": 4, "name": "恒润股份", "code": "603985", "market": "SH",
        "shares": "承立新持有1693.29万股（分3笔）",
        "platform": "阿里拍卖", "court": "江苏省江阴市人民法院",
        "date": "2026-06-13", "date_end": "2026-06-14",
        "date_display": "06-13 10:00", "date_end_display": "06-14 10:00",
        "price": "7823万/1.14亿/1.94亿", "price_note": "3笔合计3.86亿元",
        "summary": "分3笔同时拍卖：343.29万股(起拍7823万)、500万股(起拍1.1394亿)、850万股(起拍1.937亿)，均为无限售流通股，评估价70%起拍。",
        "background": "恒润股份主营法兰锻件，原控股股东承立新因金融借款合同纠纷被江阴法院强制执行。此前1280万股已分两次成交，本次为剩余持股。",
        "status": "4天后"
    },
    {
        "id": 5, "name": "ST百利", "code": "603959", "market": "SH",
        "shares": "原控股股东新海新持有2946.9万股",
        "platform": "阿里拍卖", "court": "上海市崇明区人民法院",
        "date": "2026-06-13", "date_end": "2026-06-16",
        "date_display": "06-13 10:00", "date_end_display": "06-16 10:00",
        "price": "约1.28亿元", "price_note": "占总股本6.01%",
        "summary": "拍卖原控股股东西藏新海新创业投资持有的ST百利2946.9万股无限售流通股，展示起拍价约1.28亿元，可能导致公司控制权发生变化。",
        "background": "ST百利主营工程咨询设计，原控股股东新海新因债务纠纷被强制执行。若拍卖成功，新海新持股将降至5.82%，公司控制权格局或生变。公司此前已被证监会立案。",
        "status": "4天后"
    },
    {
        "id": 6, "name": "*ST亚士", "code": "603378", "market": "SH",
        "shares": "李金钟持有200万股",
        "platform": "阿里拍卖", "court": "浙江省杭州市上城区人民法院",
        "date": "2026-06-15", "date_end": "2026-06-16",
        "date_display": "06-15 10:00", "date_end_display": "06-16 10:00",
        "price": "754.6万元", "price_note": "市场价1078万×70%",
        "summary": "拍卖被执行人李金钟持有的*ST亚士200万股无限售流通股，市场参考价1078万元，起拍价754.6万元，保证金150.92万元。",
        "background": "亚士创能(现*ST亚士)主营建筑节能材料，因经营困难被ST。此前控股股东创能明850万股于5月23日已成交。被执行人李金钟所持股份因个人债务被司法冻结并强制拍卖。",
        "status": "6天后"
    },
    {
        "id": 7, "name": "万马科技", "code": "300698", "market": "SZ",
        "shares": "闫楠持有240万股（分4笔各60万股）",
        "platform": "阿里拍卖", "court": "浙江省杭州市上城区人民法院",
        "date": "2026-06-15", "date_end": "2026-06-16",
        "date_display": "06-15 10:00", "date_end_display": "06-16 10:00",
        "price": "1584.66万元/笔", "price_note": "4笔合计6338.64万元",
        "summary": "分4笔同时拍卖被执行人闫楠持有的万马科技各60万股(合计240万股)，每笔起拍1584.66万元，市场价2263.8万/笔，保证金316.9万/笔。",
        "background": "万马科技主营智能配用电设备。被执行人闫楠所持4笔各60万股因债务纠纷被杭州上城区法院冻结并同时挂拍。此前3月首批在京东平台流拍，92万股5月完成二拍成交。",
        "status": "6天后"
    },
    {
        "id": 8, "name": "深华发A", "code": "000020", "market": "SZ",
        "shares": "武汉中恒集团持有318.99万股（分5笔，二拍）",
        "platform": "京东拍卖", "court": "广东省深圳市中级人民法院",
        "date": "2026-06-17", "date_end": "2026-06-18",
        "date_display": "06-17 10:00", "date_end_display": "06-18 10:00",
        "price": "待公告", "price_note": "一拍于5月28-29日流拍",
        "summary": "控股股东武汉中恒集团318.99万股将被第二次司法拍卖，分5笔（4次各64万股、1次62.99万股），一拍已于京东平台流拍。",
        "background": "深华发A主营精密注塑件及显示器。控股股东武汉中恒集团因与万科合同纠纷被深圳中院强制执行。一拍流拍后启动二拍程序，公司称已达成债务冲抵共识并申诉异议。",
        "status": "8天后"
    },
    {
        "id": 9, "name": "古鳌科技", "code": "300551", "market": "SZ",
        "shares": "陈崇军持有260万股（5%以上股东）",
        "platform": "阿里拍卖", "court": "湖南省长沙市望城区人民法院",
        "date": "2026-06-20", "date_end": "2026-06-21",
        "date_display": "06-20 10:00", "date_end_display": "06-21 10:00",
        "price": "待公告", "price_note": "参考前20日均价",
        "summary": "5%以上股东陈崇军260万股（占总股本0.76%）将在淘宝网司法拍卖平台拍卖。此前实控人已被判刑，持股持续被处置。",
        "background": "古鳌科技主营智能现金处理设备。第一大股东陈崇军因个人债务纠纷及操纵证券市场罪被判刑，其持股被长沙望城区法院持续强制执行。若全部成交，控制权格局生变。",
        "status": "11天后"
    },
    {
        "id": 10, "name": "劲嘉股份", "code": "002191", "market": "SZ",
        "shares": "劲嘉创投持有3500万股（控股股东）",
        "platform": "阿里拍卖", "court": "广东省深圳市中级人民法院",
        "date": "2026-06-23", "date_end": "2026-06-24",
        "date_display": "06-23 10:00", "date_end_display": "06-24 10:00",
        "price": "待公告", "price_note": "参考前20日均价",
        "summary": "控股股东劲嘉创投3500万股（占总股本2.41%）将被司法拍卖。若成交，控股股东持股比例将由22.37%降至19.96%。",
        "background": "劲嘉股份主营烟标印刷及包装，控股股东劲嘉创投因借款合同纠纷被深圳中院执行。此前已有4100万股被拍卖成交（2026年1-5月），本次为新一轮拍卖，控股股东持股持续稀释。",
        "status": "14天后"
    },
    {
        "id": 11, "name": "松井股份", "code": "688157", "market": "SH",
        "shares": "凌云剑持有56.87万股（实际控制人）",
        "platform": "京东拍卖", "court": "广东省深圳市中级人民法院",
        "date": "2026-06-25", "date_end": "2026-06-26",
        "date_display": "06-25 10:00", "date_end_display": "06-26 10:00",
        "price": "1943万元", "price_note": "占总股本0.36%",
        "summary": "实际控制人凌云剑直接持有的56.87万股将被司法拍卖，起拍价1943万元。拍卖在京东网司法拍卖平台进行。",
        "background": "松井股份主营高端消费电子涂料及汽车涂料。实控人凌云剑因与星帝企业有限公司合同纠纷被深圳中院强制执行。其直接及间接合计持有公司约7600万股，本次拍卖占其直接持股的一部分。",
        "status": "16天后"
    },
    {
        "id": 12, "name": "惠同股份", "code": "834189", "market": "新三板",
        "shares": "朱春树持有500万股",
        "platform": "阿里拍卖", "court": "浙江省丽水市遂昌县人民法院",
        "date": "2026-06-27", "date_end": "2026-06-28",
        "date_display": "06-27 10:00", "date_end_display": "06-28 10:00",
        "price": "393.16万元", "price_note": "新三板挂牌公司",
        "summary": "拍卖股东朱春树持有的惠同股份500万股，起拍价393.162万元。惠同股份为新三板挂牌公司，主营金属纤维及制品。",
        "background": "惠同股份(834189)为新三板创新层企业，主营金属纤维及制品，应用于军工、化工等领域。股东朱春树因个人法律纠纷被遂昌县法院强制执行，其持股被裁定拍卖。",
        "status": "18天后"
    },
    {
        "id": 13, "name": "迅游科技", "code": "300467", "market": "SZ",
        "shares": "袁旭持有830.54万股（创始人/第一大股东）",
        "platform": "阿里拍卖", "court": "四川省成都市中级人民法院",
        "date": "2026-06-29", "date_end": "2026-06-30",
        "date_display": "06-29 14:00", "date_end_display": "06-30 14:00",
        "price": "2.7亿元", "price_note": "市场价3.22亿×84%",
        "summary": "拍卖迅游科技创始人袁旭持有的830.54万股(占总股本约4.09%)，市场价约3.22亿元，起拍价2.7亿元，保证金1611.75万元。",
        "background": "袁旭为迅游科技创始人、第一大股东，因借款担保纠纷被成都中院强制执行。公司主营游戏加速器业务。袁旭此前多次质押股份，本次系其个人债务升级所致，若成交将显著改变持股结构。",
        "status": "20天后"
    },
    {
        "id": 14, "name": "*ST三房", "code": "600370", "market": "SH",
        "shares": "三房巷集团持有4917万股（控股股东）",
        "platform": "京东拍卖", "court": "江苏省江阴市人民法院",
        "date": "2026-06-30", "date_end": "2026-07-01",
        "date_display": "约6月底-7月初", "date_end_display": "告知书后约30日",
        "price": "约4509万元", "price_note": "按1.31元×70%估算",
        "summary": "控股股东三房巷集团4917万股将被拍卖，累计拟拍卖达3.84亿股。拍卖平台为京东网司法拍卖平台，具体时间以法院公告为准。",
        "background": "*ST三房主营化纤及薄膜，因金融借款合同纠纷被江阴法院执行。控股股东持有约73%股份，大量被轮候冻结。公司股价持续暴跌，目前收于1.31元附近，存在面值退市风险。",
        "status": "待定"
    },
    {
        "id": 15, "name": "申港证券", "code": "非上市", "market": "",
        "shares": "上海煌富贸易持有合计9846.16万股（3笔）",
        "platform": "阿里拍卖", "court": "河南省许昌市魏都区人民法院",
        "date": "2026-07-01", "date_end": "2026-07-02",
        "date_display": "07-01 10:00", "date_end_display": "07-02 10:00",
        "price": "1646万/5030万/5030万", "price_note": "3笔合计1.17亿元",
        "summary": "分3笔拍卖申港证券股权：1384.62万股(起拍1646.06万)、4230.77万股×2笔(各起拍5029.63万)，评估价即起拍价，保证金20%。",
        "background": "申港证券为非上市券商(总部上海)，被执行人上海煌富贸易所持股权因债务纠纷被许昌魏都区法院强制执行拍卖，3笔标的同日开拍。",
        "status": "22天后"
    },
    {
        "id": 16, "name": "古鳌科技", "code": "300551", "market": "SZ",
        "shares": "陈崇军持有255万股（第一大股东）",
        "platform": "阿里拍卖", "court": "湖南省长沙市望城区人民法院",
        "date": "2026-07-04", "date_end": "2026-07-05",
        "date_display": "07-04 10:00", "date_end_display": "07-05 10:00",
        "price": "待公告", "price_note": "参考前20日均价",
        "summary": "第一大股东陈崇军255万股（占总股本0.75%）将被司法拍卖。若成交，陈崇军持股将降至16.60%，实控人表决权降至21.10%。",
        "background": "古鳌科技第一大股东陈崇军因操纵证券市场罪被判刑，个人债务纠纷持续发酵。长沙望城区法院分批处置其持股，此前240万股已公告拍卖，本次为新增255万股。",
        "status": "25天后"
    }
]

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "拍卖标的跟踪"

# Title
ws.merge_cells('A1:H1')
ws['A1'] = '上市公司股票司法拍卖标的跟踪表'
ws['A1'].font = Font(name='Microsoft YaHei', size=18, bold=True, color='1E3A5F')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

# Subtitle
ws.merge_cells('A2:H2')
ws['A2'] = f'数据来源：阿里/京东司法拍卖平台、上市公司公告    更新日期：2026-06-09    筛选范围：尚未结拍'
ws['A2'].font = Font(name='Microsoft YaHei', size=10, color='8F9BB3')
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 24

# Headers
headers = ['序号', '标的名称', '处置机构', '起拍时间', '起拍价格', '公告核心要素', '挂拍背景', '状态']
header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment

ws.row_dimensions[4].height = 28

# Data rows
thin_border = Border(
    left=Side(style='thin', color='E8ECF0'),
    right=Side(style='thin', color='E8ECF0'),
    top=Side(style='thin', color='E8ECF0'),
    bottom=Side(style='thin', color='E8ECF0')
)

for row_idx, item in enumerate(auction_data, 5):
    row = [
        item['id'],
        f"{item['name']}\n{item['code']}.{item['market'] if item['market'] else '非上市'}\n{item['shares']}\n【{item['platform']}】",
        item['court'],
        f"{item['date_display']}\n至 {item['date_end_display']}",
        f"{item['price']}\n{item['price_note']}",
        item['summary'],
        item['background'],
        item['status']
    ]
    
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name='Microsoft YaHei', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = thin_border
    
    # Status color
    status_cell = ws.cell(row=row_idx, column=8)
    if item['status'] == '进行中':
        status_cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color='CF1322')
    elif '天后' in item['status'] and int(item['status'].replace('天后', '')) <= 7:
        status_cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color='D46B08')
    else:
        status_cell.font = Font(name='Microsoft YaHei', size=10, color='096DD9')
    
    ws.row_dimensions[row_idx].height = 90

# Column widths
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 38
ws.column_dimensions['G'].width = 42
ws.column_dimensions['H'].width = 10

# Freeze panes
ws.freeze_panes = 'A5'

# Add summary sheet
ws2 = wb.create_sheet("统计摘要")
ws2['A1'] = '统计项目'
ws2['B1'] = '数值'
ws2['A1'].font = Font(bold=True)
ws2['B1'].font = Font(bold=True)

summary = [
    ('跟踪标的总数', len(auction_data)),
    ('阿里拍卖平台', sum(1 for x in auction_data if x['platform'] == '阿里拍卖')),
    ('京东拍卖平台', sum(1 for x in auction_data if x['platform'] == '京东拍卖')),
    ('进行中', sum(1 for x in auction_data if x['status'] == '进行中')),
    ('7日内开拍', sum(1 for x in auction_data if '天后' in x['status'] and int(x['status'].replace('天后', '')) <= 7)),
    ('A股上市公司', sum(1 for x in auction_data if x['market'] in ['SZ', 'SH'])),
    ('新三板挂牌', sum(1 for x in auction_data if x['market'] == '新三板')),
    ('非上市券商', sum(1 for x in auction_data if x['market'] == '' and '证券' in x['name'])),
    ('更新日期', '2026-06-09'),
]

for i, (k, v) in enumerate(summary, 2):
    ws2[f'A{i}'] = k
    ws2[f'B{i}'] = v

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 15

# Save
output_path = r'C:\Users\botang\.qoderwork\workspace\mps9p9fozwjqroo2\outputs\上市公司股票司法拍卖跟踪表.xlsx'
wb.save(output_path)
print(f'Excel file saved to: {output_path}')
